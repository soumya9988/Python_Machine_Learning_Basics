import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Product Vs Price'

bold_font = Font(bold=True, size=13, italic=True)
ws['A1'].font = bold_font
ws['A1'].value = 'Product'
ws['B1'].font = bold_font
ws['B1'].value = 'Price'
ws['C1'].font = bold_font
ws['C1'].value = 'Qty'

ws.append(['Apples', 3, 1.25])
ws.append(['Bread', 6, 1])
ws.append(['Chocolates', 1.25, 1])
ws.append(['Milk', 4.50, 1])
ws.append(['Potato', 0.75, 2.25])
ws.append(['Chicken', 18, 6.5])

new_sheet = wb.create_sheet('Invert Sheet')

row_no = ws.max_row
col_no = ws.max_column
new_sheet['A1'].font = bold_font
new_sheet['A2'].font = bold_font
new_sheet['A3'].font = bold_font
for i in range(1, col_no + 1):
    for j in range(1, row_no + 1):
        new_sheet.cell(row=i, column=j).value = ws.cell(row=j, column=i).value

wb.save('excel_invert_table.xlsx')


