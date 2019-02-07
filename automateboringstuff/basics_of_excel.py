from openpyxl import Workbook, load_workbook

# To create a new work book
wb = Workbook()

# Get the active work sheet in the workbook
work_sheet = wb.active

# Populating the data in the work sheet
work_sheet['A1'] = 'Apples'
work_sheet['B1'] = 73
work_sheet['A2'] = 'Cherry'
work_sheet['B2'] = 12
work_sheet['A3'] = 'Eggplant'
work_sheet['B3'] = 38
work_sheet['A4'] = 'Grapes'
work_sheet['B4'] = 9
work_sheet.append(['Ice cream', 91])
wb.save('sample_excel.xlsx')

work_book = load_workbook('sample_excel.xlsx')
print(type(work_book))
print(work_book.sheetnames)
cur_work_sheet = work_book.get_sheet_by_name('Sheet')
print(type(cur_work_sheet))

first_row = work_sheet['A1']
print(first_row.value, first_row.coordinate, first_row.row, first_row.column)

col_size = work_sheet.max_column
row_size = work_sheet.max_row
print(col_size, row_size)

for i in range(1, row_size + 1):
    for j in range(1, col_size + 1):
        print(work_sheet.cell(row=i, column=j).coordinate, '-->', work_sheet.cell(row=i, column=j).value, end=' ')
    print('')

