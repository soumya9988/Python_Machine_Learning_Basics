import openpyxl

price_update = {'Celery': 1.19,
                'Garlic': 3.07,
                'Lemon' : 1.27
               }

work_book = openpyxl.load_workbook('produceSales.xlsx')
work_sheet = work_book.active

row_to_check = work_sheet.max_row + 1

for produce in range(2, 100):
    item = work_sheet.cell(row=produce, column= 1).value
    if item in price_update:
        work_sheet.cell(row=produce, column= 2).value = price_update[item]

work_book.save('produceSales.xlsx')




