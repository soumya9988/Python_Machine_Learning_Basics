import openpyxl

# Creating a workbook
work_book = openpyxl.Workbook()
print(work_book.sheetnames)
work_sheet = work_book.active
print(work_sheet.title)

# Renaming the work sheet
work_sheet.title = 'Bacon Egg and Spam'
print(work_sheet.title)

# Adding work sheets
work_book.create_sheet('New Bacon')
work_book.create_sheet('New Egg')
work_book.create_sheet('New Spam')

# Getting the work sheets as a list
list_ws = work_book.sheetnames
print(list_ws)

# Adding values to work sheet
work_sheet.append(['Apple', 20, 'Fresh'])
work_sheet.append(['Fish', 28, 'Frozen'])
work_sheet.append(['Strawberry', 13, 'Fresh'])

new_sheet = work_book['New Bacon']
new_sheet.append([10, 100.00])
new_sheet.append([2, 12.50])

work_book.save('basic_excel.xlsx')

# Removing worksheet
wb_1 = openpyxl.load_workbook('basic_excel.xlsx')
wb_1.remove(wb_1['New Spam'])
wb_1.save('basic_excel.xlsx')

