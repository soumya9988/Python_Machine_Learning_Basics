'''
Create a program multiplication_table.py that takes a number N from the command line
and creates an N×N multiplication table in an Excel spreadsheet.
'''

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys

if len(sys.argv) > 1:
    n = int(sys.argv[1])
else:
    n = 10

# Creating a new work book and a work sheet with name Multiplication table
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Multiplication Table'

# Printing the bold header
bold_font = Font(bold=True, size=15)
for i in range(2, n + 2):

    ws['A'+str(i)].font = bold_font
    ws['A' + str(i)].value = i - 1

    j = get_column_letter(i)
    ws[j + str(1)].font = bold_font
    ws[j + str(1)].value = i - 1

# Generating the multiplication table
for i in range(2, n + 2):
    for j in range(2, n + 2):
        ws.cell(row=i, column=j).value = (i-1) * (j-1)

# Finally saving the excel
ws.freeze_panes = 'A20'
wb.save('multiplication.xlsx')
